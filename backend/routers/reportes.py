"""
routers/reportes.py — Launch Scope: Cobranza y Reportes

Endpoints de cobranza y reportes contables del tenant.
Extraídos de routers/pagos.py en la limpieza de P2 del refactor (2026-04-07).

Rutas:
  GET /cobranza/resumen      — posición de cobranza del tenant
  GET /cobranza/vencidas     — documentos vencidos con saldo pendiente
  GET /reporte/mensual       — Excel mensual para el contador
"""

import calendar
import io
from datetime import datetime
from decimal import Decimal

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

import crud
import models
import schemas
from api_dependencies import get_current_user, get_db_tenant
from services.document_flow_service import (
    DOCUMENT_KIND_CREDIT_NOTE,
    DOCUMENT_KIND_DEBIT_NOTE,
    DOCUMENT_KIND_FISCAL_DOCUMENT,
)
from services.fiscal_balance_service import get_fiscal_document_balance
from services.quote_observation_service import observation_lines_to_plain_text

router = APIRouter(tags=["reportes"])


MONTHLY_REPORT_HEADERS = [
    "Tipo",
    "Serie",
    "Correlativo",
    "Fecha Emision",
    "Cliente",
    "RUC / DNI",
    "Tipo Doc.",
    "Signo Fiscal",
    "Gravada",
    "Exonerada",
    "Inafecta",
    "IGV",
    "Total",
    "Moneda",
    "Estado Pago",
    "Monto Pagado",
    "Saldo Pendiente",
    "Condicion Pago",
    "Observaciones",
]

MONTHLY_REPORT_MONEY_COLUMNS = [9, 10, 11, 12, 13, 16, 17]
MONTHLY_REPORT_TOTAL_COLUMNS = {
    "total_gravada": 9,
    "total_exonerada": 10,
    "total_inafecta": 11,
    "total_igv": 12,
    "total_venta": 13,
    "monto_pagado": 16,
    "saldo_pendiente": 17,
}


def _report_decimal(value) -> Decimal:
    return Decimal(str(value if value is not None else "0")).quantize(Decimal("0.01"))


def _fiscal_sign(doc) -> Decimal:
    if doc.document_kind == "credit_note" or doc.tipo_comprobante == "07":
        return Decimal("-1")
    return Decimal("1")


def _signed_fiscal_totals(doc) -> dict[str, Decimal]:
    sign = _fiscal_sign(doc)
    return {
        "total_gravada": _report_decimal(doc.total_gravada) * sign,
        "total_exonerada": _report_decimal(doc.total_exonerada) * sign,
        "total_inafecta": _report_decimal(doc.total_inafecta) * sign,
        "total_igv": _report_decimal(doc.total_igv) * sign,
        "total_venta": _report_decimal(doc.total_venta) * sign,
    }


def _monthly_collection_amounts(
    db: Session,
    tenant_id: int,
    doc,
) -> tuple[Decimal, Decimal]:
    if (
        doc.document_kind == DOCUMENT_KIND_FISCAL_DOCUMENT
        and doc.tipo_comprobante in {"01", "03"}
    ):
        balance = get_fiscal_document_balance(db, tenant_id, doc.id)
        return (
            _report_decimal(balance.payments_total),
            _report_decimal(balance.saldo_pendiente),
        )

    if doc.document_kind in {DOCUMENT_KIND_CREDIT_NOTE, DOCUMENT_KIND_DEBIT_NOTE}:
        return Decimal("0.00"), Decimal("0.00")

    return _report_decimal(doc.monto_pagado), _report_decimal(doc.saldo_pendiente)


def _monthly_payment_status(doc, monto_pagado: Decimal, saldo: Decimal) -> str:
    if doc.document_kind in {DOCUMENT_KIND_CREDIT_NOTE, DOCUMENT_KIND_DEBIT_NOTE}:
        return "ajuste"
    if saldo <= 0 and monto_pagado > 0:
        return "pagado"
    if monto_pagado > 0:
        return "parcial"
    return doc.payment_status if hasattr(doc, "payment_status") else ""


@router.get(
    "/cobranza/resumen",
    response_model=schemas.CobranzaResumenResponse,
    summary="Resumen de cobranza del tenant",
)
def cobranza_resumen(
    db: Session = Depends(get_db_tenant),
    current_user: models.User = Depends(get_current_user),
):
    """
    Devuelve un resumen de la posición de cobranza:
    - total por cobrar
    - total vencido
    - total cobrado en el mes
    - contadores de documentos por estado
    """
    data = crud.get_cobranza_resumen(db, current_user.tenant_id)
    return schemas.CobranzaResumenResponse(**data)


@router.get(
    "/cobranza/vencidas",
    response_model=list[schemas.CobranzaVencidaItem],
    summary="Lista de documentos vencidos con saldo pendiente",
)
def cobranza_vencidas(
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=25, ge=1, le=100),
    q: str | None = Query(default=None, max_length=120),
    scope: str = Query(default="overdue", pattern="^(overdue|active)$"),
    db: Session = Depends(get_db_tenant),
    current_user: models.User = Depends(get_current_user),
):
    """
    Lista cotizaciones cuya fecha_vencimiento ya pasó y aún tienen saldo pendiente.
    Ordena de la más antigua (mayor riesgo) a la más reciente.
    """
    ahora = datetime.now()
    try:
        cotizaciones = crud.get_cobranza_vencida(
            db,
            current_user.tenant_id,
            skip=skip,
            limit=limit,
            q=q,
            scope=scope,
        )
    except ValueError as exc:
        raise HTTPException(400, str(exc))

    result = []
    for cot in cotizaciones:
        dias_vencido = 0
        if cot.fecha_vencimiento:
            dias_vencido = (ahora.date() - cot.fecha_vencimiento.date()).days

        cliente_nombre = cot.cliente_nombre or cot.cliente_nombre_alt or ""
        document_number = None
        if cot.serie and cot.correlativo is not None:
            document_number = f"{cot.serie}-{str(cot.correlativo).zfill(6)}"

        total_venta = cot.total_venta or Decimal("0")
        monto_pagado = cot.monto_pagado or Decimal("0")
        if monto_pagado >= total_venta and total_venta > 0:
            payment_status = "pagado"
        elif dias_vencido > 0:
            payment_status = "vencido"
        elif monto_pagado > 0:
            payment_status = "parcial"
        else:
            payment_status = "pendiente"

        result.append({
            "id": cot.id,
            "cotizacion_id": cot.id,
            "serie": cot.serie,
            "correlativo": cot.correlativo,
            "document_number": document_number,
            "internal_order_number": cot.internal_order_number,
            "cliente_nombre": cliente_nombre,
            "cliente_documento": cot.cliente_documento or "",
            "cliente": {
                "razon_social": cliente_nombre,
                "numero_documento": cot.cliente_documento or "",
            },
            "document_kind": cot.document_kind,
            "estado": cot.estado,
            "fecha_emision": cot.fecha_emision,
            "fecha_vencimiento": cot.fecha_vencimiento,
            "moneda": cot.moneda or "PEN",
            "total_venta": total_venta,
            "monto_pagado": monto_pagado,
            "saldo_pendiente": cot.saldo_pendiente or Decimal("0"),
            "dias_vencido": dias_vencido,
            "payment_status": payment_status,
        })

    return result


@router.get(
    "/reporte/mensual",
    summary="Reporte mensual de documentos fiscales (Excel para contador)",
)
def reporte_mensual_excel(
    anio: int = Query(..., ge=2020, le=2100, description="Año del reporte"),
    mes: int = Query(..., ge=1, le=12, description="Mes del reporte (1-12)"),
    db: Session = Depends(get_db_tenant),
    current_user: models.User = Depends(get_current_user),
):
    """
    Genera un archivo Excel con todos los comprobantes fiscales emitidos en el mes.
    Diseñado para ser entregado directamente al contador.
    Incluye: número de comprobante, fecha, cliente, RUC, base imponible, IGV, total,
    estado de pago, monto pagado, saldo pendiente y condición de pago.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl no está instalado en el servidor.")

    docs = crud.get_reporte_mensual(db, current_user.tenant_id, anio, mes)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Reporte {mes:02d}-{anio}"

    # ── Cabecera empresa ──────────────────────────────────────────────────────
    tenant = (
        db.query(models.Tenant)
        .filter(models.Tenant.id == current_user.tenant_id)
        .first()
    )
    empresa = tenant.business_name if tenant else "Empresa"
    ruc_empresa = tenant.business_ruc if tenant else ""

    last_report_col = get_column_letter(len(MONTHLY_REPORT_HEADERS))

    ws.merge_cells(f"A1:{last_report_col}1")
    ws["A1"] = f"REPORTE MENSUAL DE COMPROBANTES — {empresa} (RUC {ruc_empresa})"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells(f"A2:{last_report_col}2")
    nombre_mes = calendar.month_name[mes].upper()
    ws["A2"] = f"PERIODO: {nombre_mes} {anio}"
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.append([])  # fila vacía

    # ── Encabezados de tabla ──────────────────────────────────────────────────
    header_row = 4
    ws.append(MONTHLY_REPORT_HEADERS)

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, _ in enumerate(MONTHLY_REPORT_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # ── Filas de datos ────────────────────────────────────────────────────────
    TIPO_LABEL = {"01": "Factura", "03": "Boleta", "07": "Nota Crédito", "08": "Nota Débito"}
    totales_gravada = Decimal("0")
    totales_exonerada = Decimal("0")
    totales_inafecta = Decimal("0")
    totales_igv = Decimal("0")
    totales_total = Decimal("0")
    totales_pagado = Decimal("0")
    totales_saldo = Decimal("0")

    alt_fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")

    for row_num, doc in enumerate(docs, start=1):
        cliente = doc.cliente
        fiscal_totals = _signed_fiscal_totals(doc)
        fiscal_sign = _fiscal_sign(doc)
        monto_pagado, saldo = _monthly_collection_amounts(
            db,
            current_user.tenant_id,
            doc,
        )

        tipo_label = TIPO_LABEL.get(doc.tipo_comprobante or "", doc.tipo_comprobante or "")
        fecha_str = doc.fecha_emision.strftime("%d/%m/%Y") if doc.fecha_emision else ""
        payment_status = _monthly_payment_status(doc, monto_pagado, saldo)

        row_data = [
            tipo_label,
            doc.serie or "",
            str(doc.correlativo or "").zfill(6),
            fecha_str,
            cliente.razon_social if cliente else "",
            cliente.numero_documento if cliente else "",
            cliente.tipo_documento if cliente else "",
            int(fiscal_sign),
            float(fiscal_totals["total_gravada"]),
            float(fiscal_totals["total_exonerada"]),
            float(fiscal_totals["total_inafecta"]),
            float(fiscal_totals["total_igv"]),
            float(fiscal_totals["total_venta"]),
            doc.moneda or "PEN",
            payment_status,
            float(monto_pagado),
            float(saldo),
            doc.condicion_pago or "",
            observation_lines_to_plain_text(doc.observaciones),
        ]

        ws.append(row_data)
        data_row = header_row + row_num

        # Filas alternadas
        if row_num % 2 == 0:
            for col_idx in range(1, len(MONTHLY_REPORT_HEADERS) + 1):
                ws.cell(row=data_row, column=col_idx).fill = alt_fill

        # Formato numerico para columnas monetarias.
        for col_idx in MONTHLY_REPORT_MONEY_COLUMNS:
            ws.cell(row=data_row, column=col_idx).number_format = '#,##0.00'

        totales_gravada += fiscal_totals["total_gravada"]
        totales_exonerada += fiscal_totals["total_exonerada"]
        totales_inafecta += fiscal_totals["total_inafecta"]
        totales_igv += fiscal_totals["total_igv"]
        totales_total += fiscal_totals["total_venta"]
        totales_pagado += monto_pagado
        totales_saldo += saldo

    # ── Fila de totales ───────────────────────────────────────────────────────
    total_row = header_row + len(docs) + 1
    ws.cell(row=total_row, column=1, value="TOTALES")
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    totales = {
        MONTHLY_REPORT_TOTAL_COLUMNS["total_gravada"]: float(totales_gravada),
        MONTHLY_REPORT_TOTAL_COLUMNS["total_exonerada"]: float(totales_exonerada),
        MONTHLY_REPORT_TOTAL_COLUMNS["total_inafecta"]: float(totales_inafecta),
        MONTHLY_REPORT_TOTAL_COLUMNS["total_igv"]: float(totales_igv),
        MONTHLY_REPORT_TOTAL_COLUMNS["total_venta"]: float(totales_total),
        MONTHLY_REPORT_TOTAL_COLUMNS["monto_pagado"]: float(totales_pagado),
        MONTHLY_REPORT_TOTAL_COLUMNS["saldo_pendiente"]: float(totales_saldo),
    }
    for col_idx, val in totales.items():
        cell = ws.cell(row=total_row, column=col_idx, value=val)
        cell.number_format = '#,##0.00'
        cell.font = Font(bold=True)

    # ── Ancho de columnas ─────────────────────────────────────────────────────
    col_widths = [12, 8, 12, 14, 35, 14, 10, 12, 14, 14, 14, 12, 14, 8, 14, 14, 16, 16, 30]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ── Resumen final ─────────────────────────────────────────────────────────
    summary_start = total_row + 3
    ws.cell(row=summary_start, column=1, value="RESUMEN DEL PERIODO").font = Font(bold=True, size=11)
    ws.cell(row=summary_start + 1, column=1, value="Total documentos emitidos:")
    ws.cell(row=summary_start + 1, column=2, value=len(docs))
    ws.cell(row=summary_start + 2, column=1, value="Total facturado:")
    ws.cell(row=summary_start + 2, column=2, value=float(totales_total)).number_format = '#,##0.00'
    ws.cell(row=summary_start + 2, column=2).number_format = '#,##0.00'
    ws.cell(row=summary_start + 3, column=1, value="Total cobrado:")
    ws.cell(row=summary_start + 3, column=2, value=float(totales_pagado))
    ws.cell(row=summary_start + 3, column=2).number_format = '#,##0.00'
    ws.cell(row=summary_start + 4, column=1, value="Saldo pendiente:")
    ws.cell(row=summary_start + 4, column=2, value=float(totales_saldo))
    ws.cell(row=summary_start + 4, column=2).number_format = '#,##0.00'

    # ── Serializar y devolver ─────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"reporte_{anio}_{mes:02d}_{current_user.tenant_id}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
